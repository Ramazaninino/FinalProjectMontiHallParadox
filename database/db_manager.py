import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)


class DatabaseManager:
    """
    Manages all SQLite3 database operations.
    Demonstrates: sqlite3, SQL SELECT, SQL JOIN
    """

    DB_PATH = "monty_hall.db"
    EXPORTS_DIR = "exports"

    def __init__(self):
        self._connection: sqlite3.Connection | None = None
        self._connect()
        self._create_tables()

    def __str__(self):
        return f"DatabaseManager(db='{self.DB_PATH}')"

    def __repr__(self):
        return f"DatabaseManager(path='{self.DB_PATH}', connected={self._connection is not None})"

    def __del__(self):
        self.close()

    # Connection management
    def _connect(self):
        self._connection = sqlite3.connect(self.DB_PATH)
        self._connection.row_factory = sqlite3.Row  # Dict-like rows
        self._connection.execute("PRAGMA foreign_keys = ON")

    def close(self):
        if self._connection:
            self._connection.close()
            self._connection = None

    def _create_tables(self):
        """Create all required tables if they don't exist."""
        cursor = self._connection.cursor()

        # Main game sessions table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS game_sessions (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT    NOT NULL,
                door_count  INTEGER NOT NULL,
                mode        TEXT    NOT NULL,
                total_games INTEGER DEFAULT 0,
                wins        INTEGER DEFAULT 0,
                num_games   INTEGER DEFAULT 0,
                created_at  TEXT    DEFAULT (strftime('%d.%m.%Y %H:%M', 'now', 'localtime'))
            )
        """)

        # Individual rounds table (for detailed history)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS game_rounds (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id   INTEGER NOT NULL,
                round_number INTEGER NOT NULL,
                switched     INTEGER NOT NULL,
                won          INTEGER NOT NULL,
                created_at   TEXT DEFAULT (strftime('%d.%m.%Y %H:%M', 'now', 'localtime')),
                FOREIGN KEY (session_id) REFERENCES game_sessions(id) ON DELETE CASCADE
            )
        """)

        self._connection.commit()

    # Session CRUD
    def create_session(self, name: str, door_count: int, mode: str,
                       num_games: int = 0) -> int:
        """Create a new game session. Returns the new session ID."""
        cursor = self._connection.cursor()
        cursor.execute("""
            INSERT INTO game_sessions (name, door_count, mode, num_games)
            VALUES (?, ?, ?, ?)
        """, (name, door_count, mode, num_games))
        self._connection.commit()
        return cursor.lastrowid

    def update_session_stats(self, session_id: int, total_games: int, wins: int):
        """Update win/loss stats for a session."""
        self._connection.execute("""
            UPDATE game_sessions
            SET total_games = ?, wins = ?
            WHERE id = ?
        """, (total_games, wins, session_id))
        self._connection.commit()

    def add_round(self, session_id: int, round_number: int,
                  switched: bool, won: bool):
        """Record an individual game round."""
        self._connection.execute("""
            INSERT INTO game_rounds (session_id, round_number, switched, won)
            VALUES (?, ?, ?, ?)
        """, (session_id, round_number, int(switched), int(won)))
        self._connection.commit()

    def get_all_sessions(self) -> list[dict]:
        """
        Fetch all sessions with aggregated stats.
        Demonstrates: SQL SELECT with LEFT JOIN and GROUP BY
        """
        cursor = self._connection.cursor()
        cursor.execute("""
            SELECT
                gs.id,
                gs.name,
                gs.door_count,
                gs.mode,
                gs.total_games,
                gs.wins,
                gs.num_games,
                gs.created_at,
                ROUND(
                    CASE WHEN gs.total_games > 0
                         THEN CAST(gs.wins AS REAL) / gs.total_games * 100
                         ELSE 0
                    END, 1
                ) AS win_rate,
                COUNT(gr.id) AS round_count
            FROM game_sessions gs
            LEFT JOIN game_rounds gr ON gs.id = gr.session_id
            GROUP BY gs.id
            ORDER BY gs.created_at DESC
        """)
        return [dict(row) for row in cursor.fetchall()]

    def get_session_by_id(self, session_id: int) -> dict | None:
        """Get a single session with its rounds count."""
        cursor = self._connection.cursor()
        cursor.execute("""
            SELECT
                gs.*,
                COUNT(gr.id) AS round_count,
                ROUND(
                    CASE WHEN gs.total_games > 0
                         THEN CAST(gs.wins AS REAL) / gs.total_games * 100
                         ELSE 0
                    END, 1
                ) AS win_rate
            FROM game_sessions gs
            LEFT JOIN game_rounds gr ON gs.id = gr.session_id
            WHERE gs.id = ?
            GROUP BY gs.id
        """, (session_id,))
        row = cursor.fetchone()
        return dict(row) if row else None

    def get_rounds_for_session(self, session_id: int) -> list[dict]:
        """Get all rounds for a specific session."""
        cursor = self._connection.cursor()
        cursor.execute("""
            SELECT * FROM game_rounds
            WHERE session_id = ?
            ORDER BY round_number ASC
        """, (session_id,))
        return [dict(row) for row in cursor.fetchall()]

    def delete_session(self, session_id: int):
        """Delete a session and all its rounds (cascade)."""
        self._connection.execute(
            "DELETE FROM game_sessions WHERE id = ?", (session_id,)
        )
        self._connection.commit()

    def get_global_stats(self) -> dict:
        """
        Return global statistics across all sessions.
        Uses subquery to avoid double-counting from JOIN.
        Demonstrates: SQL aggregate functions, subquery, JOIN
        """
        cursor = self._connection.cursor()
        cursor.execute("""
            SELECT
                COUNT(gs.id)                                    AS total_sessions,
                COALESCE(SUM(gs.total_games), 0)                AS total_games,
                COALESCE(SUM(gs.wins), 0)                       AS total_wins,
                ROUND(
                    CASE WHEN SUM(gs.total_games) > 0
                         THEN CAST(SUM(gs.wins) AS REAL) / SUM(gs.total_games) * 100
                         ELSE 0
                    END, 1
                )                                               AS overall_win_rate,
                COALESCE(
                    (SELECT COUNT(*) FROM game_rounds), 0
                )                                               AS total_rounds
            FROM game_sessions gs
        """)
        row = cursor.fetchone()
        return dict(row) if row else {}

    # ── Excel helpers ──────────────────────────────────────────────────────
    @staticmethod
    def _hdr_fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    @staticmethod
    def _thin_border() -> Border:
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def _set_col_widths(ws, widths: list):
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_title(self, ws, text: str, ncols: int):
        """Write a merged title row at the top of the sheet."""
        ws.append([text])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        cell.fill = self._hdr_fill("EB1D49")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(
            start_row=ws.max_row, start_column=1,
            end_row=ws.max_row, end_column=ncols
        )
        ws.row_dimensions[ws.max_row].height = 24

    def _write_header_row(self, ws, headers: list):
        """Write a styled header row."""
        ws.append(headers)
        row = ws.max_row
        border = self._hdr_fill("1A1A1A")
        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col)
            c.font = Font(name="Calibri", bold=True, color="FFFFFF")
            c.fill = border
            c.alignment = Alignment(horizontal="center")
            c.border = self._thin_border()
        ws.row_dimensions[row].height = 18

    def _write_data_row(self, ws, values: list, row_idx: int):
        """Write a data row with alternating fill."""
        ws.append(values)
        row = ws.max_row
        fill = self._hdr_fill("111111") if row_idx % 2 == 0 else self._hdr_fill("1A1A1A")
        border = self._thin_border()
        for col in range(1, len(values) + 1):
            c = ws.cell(row=row, column=col)
            c.font = Font(name="Calibri", color="FFFFFF")
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
            c.border = border

    # Excel Export (Week 13: Data Formats — Excel)
    def export_to_excel(self, session_id: int | None = None,
                        filepath: str | None = None) -> str:
        """
        Export sessions (or a specific session) to a formatted Excel file.
        Demonstrates: Excel data format handling (openpyxl)
        """
        if not filepath:
            os.makedirs(self.EXPORTS_DIR, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            name = f"session_{session_id}_{timestamp}.xlsx" if session_id \
                else f"all_sessions_{timestamp}.xlsx"
            filepath = os.path.join(self.EXPORTS_DIR, name)

        if session_id:
            sessions = [self.get_session_by_id(session_id)]
            rounds = self.get_rounds_for_session(session_id)
        else:
            sessions = self.get_all_sessions()
            rounds = []

        filename = filepath

        wb = Workbook()
        wb.active.title = "Sessions"
        ws_s = wb.active
        ws_s.sheet_view.showGridLines = False

        # ── Sheet 1: Sessions ────────────────────────────────────────────
        SESSION_COLS = ["ID", "Name", "Doors", "Mode",
                        "Total Games", "Wins", "Win Rate %", "Created At"]
        self._write_title(ws_s, "Monty Hall — Game Sessions", len(SESSION_COLS))
        ws_s.append([])  # spacer
        self._write_header_row(ws_s, SESSION_COLS)

        for idx, s in enumerate(sessions):
            if not s:
                continue
            self._write_data_row(ws_s, [
                s["id"],
                s["name"],
                s["door_count"],
                s["mode"].capitalize(),
                s["total_games"],
                s["wins"],
                float(s.get("win_rate") or 0),
                s["created_at"],
            ], idx)

        self._set_col_widths(ws_s, [6, 22, 8, 10, 14, 8, 14, 18])

        # ── Sheet 2: Rounds (only when exporting single session) ─────────
        if rounds:
            ws_r = wb.create_sheet(title="Rounds")
            ws_r.sheet_view.showGridLines = False

            session = sessions[0] if sessions else {}
            title_text = f"Rounds — {session.get('name', '')} (ID {session.get('id', '')})"
            ROUND_COLS = ["Round #", "Switched", "Won"]
            self._write_title(ws_r, title_text, len(ROUND_COLS))
            ws_r.append([])
            self._write_header_row(ws_r, ROUND_COLS)

            wins = losses = switched = stayed = 0
            for idx, r in enumerate(rounds):
                sw = bool(r["switched"])
                won = bool(r["won"])
                switched += sw
                stayed += not sw
                wins += won
                losses += not won
                self._write_data_row(ws_r, [
                    r["round_number"],
                    "Yes" if sw else "No",
                    "Yes" if won else "No",
                ], idx)

            # Summary block
            ws_r.append([])
            self._write_title(ws_r, "Summary", len(ROUND_COLS))
            summary_rows = [
                ["Total rounds", len(rounds), ""],
                ["Wins", wins, f"{wins / len(rounds) * 100:.1f}%"],
                ["Losses", losses, f"{losses / len(rounds) * 100:.1f}%"],
                ["Switched doors", switched, ""],
                ["Stayed", stayed, ""],
            ]
            for idx, row in enumerate(summary_rows):
                self._write_data_row(ws_r, row, idx)

            self._set_col_widths(ws_r, [10, 12, 10])

        # ── Sheet 3: Global Stats (only for full export) ─────────────────
        if not session_id:
            ws_g = wb.create_sheet(title="Global Stats")
            ws_g.sheet_view.showGridLines = False
            stats = self.get_global_stats()

            STATS_COLS = ["Metric", "Value"]
            self._write_title(ws_g, "Monty Hall — Overall Statistics", len(STATS_COLS))
            ws_g.append([])
            self._write_header_row(ws_g, STATS_COLS)

            stat_rows = [
                ("Total game sessions", stats.get("total_sessions", 0)),
                ("Total games played", stats.get("total_games", 0)),
                ("Total wins", stats.get("total_wins", 0)),
                ("Overall win rate", f"{stats.get('overall_win_rate', 0):.1f}%"),
                ("Total rounds recorded", stats.get("total_rounds", 0)),
            ]
            for idx, (metric, value) in enumerate(stat_rows):
                self._write_data_row(ws_g, [metric, value], idx)

            self._set_col_widths(ws_g, [28, 16])

        wb.save(filename)
        return filename
